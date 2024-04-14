import os,fitz
import openpyxl as px


def FindPDF(path):  # 指定path内にある全PDF探索
    PDF_path = []
    for root, dirs, files in os.walk(path):
        for filename in files:
            if filename.endswith('.pdf'):
               PDF_path.append(root+"/"+filename)
    return PDF_path


def Sort(result):
    result =  sorted(result)
    for i in range(len(result)):
        a = result[0]
        print(set(result))

try:
    #------ 1.Input読込
    PATH = os.getcwd()       #os.chdir('./Input')
    wb    = px.load_workbook(PATH+"/Input/Input.xlsx")
    ws    = wb.active
    for i in range(30):
        if ws.cell(row=1,column=i+1).value == "Find Word"       : idx_Word   = i+1
        if ws.cell(row=1,column=i+1).value == "Color"           : idx_Color  = i+1
        if ws.cell(row=1,column=i+1).value == "Output PDF"      : idx_PDFOUT = i+1
        if ws.cell(row=1,column=i+1).value == "Google Translate": idx_Google = i+1
        if ws.cell(row=1,column=i+1).value == "Sentence Number" : idx_SenNum = i+1
        if ws.cell(row=1,column=i+1).value == "Find Header"     : idx_Header = i+1
        if ws.cell(row=1,column=i+1).value == "Find Footer"     : idx_footer = i+1
        if ws.cell(row=1,column=i+1).value == "Find Title"      : idx_Title  = i+1

    WORDS = [str(ws.cell(row=3+i,column=idx_Word).value)  for i in range(10000)  if str(ws.cell(row=3+i,column=idx_Word).value)!= "None"]
    COLOR = [str(ws.cell(row=3+i,column=idx_Color).value) for i in range(10000)  if str(ws.cell(row=3+i,column=idx_Word).value)!= "None"]
    PDFOUT = ws.cell(row=3,column=idx_PDFOUT).value
    PDF_path = FindPDF(PATH+"/Input")    # ディレクト内の全PDF
      
    #----- 2. PDF出力
    os.makedirs(PATH+"/Output" , exist_ok=True)
    if PDFOUT == 1:
        file = open(PATH+'/Output/Output.txt', 'w')
        for pdf_path in PDF_path:
            pdf = os.path.basename(pdf_path)
            doc = fitz.open(pdf_path)
            file.write(pdf+"\n")                               # .txtに出力
            result = []        
            print("%s outputting..."%(pdf))
            for i in range(doc.page_count):
                page = doc[i]
                for m in range(len(WORDS)):
                    pos = page.search_for(WORDS[m], quads=True)
                    if len(pos)>0:                        
                        highlight = page.add_highlight_annot(pos)
                        highlight.set_colors(stroke=(fitz.pdfcolor[COLOR[m]]))
                        highlight.update()                        
                        result.append([WORDS[m],i+1])

            result = sorted(result, key=lambda x: x[0])        # アルファベット順
            for m in result:
                    file.write(" "+m[0]+",　P"+str(m[1])+"\n") # .txtに出力
            file.write("\n")                                   # .txtに出力
            doc.save(PATH+"/Output/"+pdf)            
            print("%s Done\n"%(pdf))            
        file.close()


except Exception as e:
    import traceback
    with open(PATH+'/error.log', 'a') as f:
        traceback.print_exc( file=f)
 

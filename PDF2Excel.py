import os,re,fitz
import openpyxl as px
from googletrans import Translator
import xlsxwriter

def SearchWord(TEXTS,WORDS,SPLIT,RANGE,GOOGLE):
    hitword, hittext1, hittext2 ,hittext_tra  = [],[],[],[]
    TEXTS = re.split(SPLIT, TEXTS)
    for i in range(len(TEXTS)):
        for word in WORDS:
            if  word.lower() in TEXTS[i].lower():
                hitpos1 = re.search(word, TEXTS[i], flags=re.IGNORECASE).start() # 大文字、小文字区別せず判定
                hitpos2 = re.search(word, TEXTS[i], flags=re.IGNORECASE).end()
                hitword.append(word)
                if RANGE == 0:
                    hittext1.append(TEXTS[i][:hitpos1])
                    hittext2.append(TEXTS[i][hitpos2:])

                if RANGE == 1:
                    temp1 = ''.join(TEXTS[i-RANGE]+TEXTS[i][:hitpos1])
                    temp2 = ''.join(TEXTS[i][hitpos2:]+TEXTS[i+RANGE])
                    hittext1.append(temp1)
                    hittext2.append(temp2)

                if RANGE > 1:
                    temp1 = ''.join(TEXTS[i-RANGE:i]+list(TEXTS[i][:hitpos1]))
                    temp2 = ''.join(list(TEXTS[i][hitpos2:])+TEXTS[i+1:i+RANGE+1])
                    hittext1.append(temp1)
                    hittext2.append(temp2)

                if GOOGLE == 1 :
                    temp = hittext1[-1]+hitword[-1]+hittext2[-1]
                    hittext_tra.append(translator.translate(temp, src='en', dest='ja').text)

                if GOOGLE == 0 :
                    hittext_tra.append(["google翻訳 未実施"])

    hittext1 = [ " " if i=="" else i  for i in hittext1 ] # 例外処理 空白のみの場合xlsxwriterでエラー
    hittext2 = [ " " if i=="" else i  for i in hittext2 ]
    return hitword,hittext1,hittext2,hittext_tra




#------ 1.input読込
wb    = px.load_workbook("input.xlsx")
ws    = wb.active

SPLIT =  ws["B2"].value
PATH   = ws["C2"].value
GOOGLE = ws["D2"].value
RANGE  = ws["E2"].value

WORDS = [str(ws.cell(row=2+i,column=1).value) for i in range(10000) ]
WORDS = [i for i in WORDS if i != "None"]
print("input.xlsx読込 Done/n")


#------ 2.PDF読込
os.chdir(PATH)
PDF = []
for filename in os.listdir(PATH):
    if filename.endswith('.pdf'):
        PDF.append(filename)

 
contents = []
translator = Translator()

for pdf in PDF:
    doc = fitz.open(pdf)
    print(pdf,"読込 Loading...")
    for page in range(len(doc)): # PDFを1ページずつテキスト取得
       print("　page",page,"Loading...") 
       TEXTS = doc[page].get_text('text')
       TEXTS = ([str(i) for i in TEXTS if i!="\n" ])       
       TEXTS = ''.join(TEXTS)
       hitword,hittext1,hittext2,hittext_tra = SearchWord(TEXTS,WORDS,SPLIT,RANGE,GOOGLE)
       for i in range(len(hitword)):
           contents.append([hitword[i],pdf,page+1,hittext1[i],hittext2[i],hittext_tra[i]])
    print(pdf,"読込 Done...\n")

 
#------ 3.エクセル作成
print("Excel outputing...") 
wb = xlsxwriter.Workbook('Output.xlsx')
ws = wb.add_worksheet()

#----- 4.ヘッダー出力
headers = ['No.', '検索単語','PDF名', 'ページ','本文','本文(翻訳)']
ws.write_row('A1', headers, wb.add_format({'pattern': 1, 'bg_color': 'ccff00', 'border': 1}))
ws.set_column('C:C', 20)
ws.set_column('E:F', 50)

#----- 5.検索結果出力
for y in range(len(contents)):
    ws.write(y+1,0,y+1,wb.add_format({'border': 1}))  # NO,
    for x in range(len(contents[y])):
        if x == 0:  # 検索単語
            ws.write(y+1,1, contents[y][0],wb.add_format({'border': 1})) 
        if x == 1:  # PDF名
            ws.write(y+1,2, contents[y][1],wb.add_format({'border': 1,'text_wrap': True})) 
        if x == 2:  #ページ
            ws.write(y+1,3, contents[y][2],wb.add_format({'border': 1}))
        if x == 3:  # 本文
            ws.write_rich_string(y+1,4,wb.add_format({'color': 'black'}),contents[y][3],
                                       wb.add_format({'color': 'red'}),contents[y][0],
                                       wb.add_format({'color': 'black'}),contents[y][4],
                                       wb.add_format({'text_wrap': True, 'border': 1}))            
        if x == 5: # 本文(翻訳)
            ws.write(y+1,5, contents[y][x],wb.add_format({'border': 1, 'text_wrap': True}))
wb.close()
print("Excel output Done... \n") 




#----- 6.PDF出力
os.makedirs(PATH+"/output" , exist_ok=True)
print("PDF outputing...")
for pdf in PDF:
    doc = fitz.open(pdf)
    for i in range(doc.page_count):
        for word in WORDS:
            pos = doc[i].search_for(word, quads=True)
            doc[i].add_squiggly_annot(pos)
            doc[i].add_highlight_annot(pos)

    doc.save(f'./output/{pdf}_output.pdf')
print("PDF output Done")

